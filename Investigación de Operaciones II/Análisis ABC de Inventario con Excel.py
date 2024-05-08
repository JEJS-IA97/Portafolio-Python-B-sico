import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from tabulate import tabulate

archivo = input("ingrese Archivo:")

excel_dataframe = openpyxl.load_workbook(f'{archivo}')
dataFrame = excel_dataframe.active


data = []

for row in range(1, dataFrame.max_row):
    _row = []
    for col in dataFrame.iter_cols(1, dataFrame.max_column):
        _row.append(col[row].value)
    data.append(_row)

def preguntar_agregar_articulo():
    respuesta = input("¿Desea agregar otro artículo? (responda sí o no): ")
    if respuesta.lower() == "sí":
        return True
    elif respuesta.lower() == "no":
        return False
    else:
        print("Respuesta inválida. Por favor, responda sí o no.")
        return preguntar_agregar_articulo()
#Borrar arriba

i = True
while (i == True):
    i = preguntar_agregar_articulo()
    if (i):
        print("Ingrese los datos del nuevo articulo")
        new_Precio = print("Precio Unitario: ")
        new_Cantidad = print("Cantidad:")



for i, row in enumerate(data):
    Valor_Uso = row[1] * row[2]
    valor_Uso_Total = + Valor_Uso
    data[i].append(Valor_Uso)
# Ordenamos de mayor a menor los valores segun el Valor de Uso
data = sorted(data, key=lambda x: x[3], reverse=True)

# print(valor_Uso_Total)
valor_Uso_Total = sum(row[3] for row in data)
# print(valor_Uso_Total)
for i, row in enumerate(data):
    uso_anual = round((row[3] * (100 / valor_Uso_Total)), 2)
    data[i].append(uso_anual)
# print(tabulate(data))
acumulado = 0
for row in data:
    acumulado += row[4]
    row.append(acumulado)

# print(tabulate(data))

clasificacion_ABC = ""
acum = 0
for row in data:
    acum += row[4]
    if acum <= 80:
        row.append("A")
    elif (acum > 80 and acum <= 95):
        row.append("B")
    else:
        row.append("C")


headers = ["Producto", "Uso anual", "Costo Unitario", "Valor Uso",
           "Uso Anual", "Suma acumulativa", "Clasificación ABC"]
print(tabulate(data, headers=headers, tablefmt='fancy_grid'))


nuevo_libro = openpyxl.Workbook()
hoja_calculo = nuevo_libro.active

# Escribir los encabezados en la hoja de cálculo
headers = ["Producto", "Uso anual", "Costo Unitario", "Valor_Uso",
           "uso_anual", "Suma acumulativa", "Clasificación ABC"]
for col_num, header in enumerate(headers, 1):
    col_letra = get_column_letter(col_num)
    celda = hoja_calculo[f"{col_letra}1"]
    celda.value = header
    celda.alignment = Alignment(horizontal='center')


# Escribir los datos en la hoja de cálculo
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        col_letra = get_column_letter(col_num)
        celda = hoja_calculo[f"{col_letra}{row_num}"]
        celda.value = cell_value
        celda.alignment = Alignment(horizontal='center')

# Guardar el archivo Excel
nuevo_libro.save('ABC_resultados.xlsx')
