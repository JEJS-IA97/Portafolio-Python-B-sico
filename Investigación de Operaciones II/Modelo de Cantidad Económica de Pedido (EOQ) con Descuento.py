import openpyxl
import math
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Abre el archivo Excel
try:
    workbook = openpyxl.load_workbook('EOQDescuento.xlsx')
except FileNotFoundError:
    print("No se pudo abrir el archivo")
    exit()

# Obtiene la hoja activa
worksheet = workbook.active

# Lee los datos del archivo
data = []
for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, values_only=True):
    data.append(row)

# Convierte los valores de la columna 6 a números
for row in data:
    row[6] = float(row[6])

# Calcula el costo de ordenar y almacenar para cada artículo
for row in data:
    row[6] = row[3] * (row[6] / 100)

# Encuentra el punto óptimo
a = 1
for i, row in enumerate(data):
    EOQ = math.sqrt((2 * row[4] * row[5]) / row[6])
    row.append(EOQ)
    if (row[2] is None):
        if (EOQ > row[3]):
            a = i
    elif (row[2] <= EOQ and row[3] > EOQ):
        a = i
        break  # óptimo

# Calcula el costo total para cada artículo
acum = float('inf')
for i, row in enumerate(data):
    if (i <= a):
        TCU = ((row[4] * row[5]) / row[7]) + (row[6] * (row[7] / 2))
        row.append(TCU)
        if (acum > TCU):
            acum = TCU

# Imprime los resultados
print(data)
print(f"El valor óptimo es: {acum}")
print(f"El número de cajas a pedir es: {data[a][7]}")
