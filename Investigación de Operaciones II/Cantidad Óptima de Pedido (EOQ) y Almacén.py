import openpyxl
import math
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from tabulate import tabulate

excel_dataframe = openpyxl.load_workbook('EOQ_Almacen - copia.xlsx')
dataFrame = excel_dataframe.active

data = []
# extraemos los datos del excel
for row in range(1, dataFrame.max_row):
    _row = []
    for col in dataFrame.iter_cols(1, dataFrame.max_column):
        _row.append(col[row].value)
    data.append(_row)
Almacen = float(input("Ingrese la dimension del almacen: "))

print("Procedemos a calcuar")
print("-----------------")
rows = 0
acum = 0
alfa = 0
for i, row in enumerate(data):

    y = round(math.sqrt((2 * row[1] * row[2]) / (2 * row[3])),2)
    print(f"K={row[1]}, D={row[2]}, H={row[3]}")
    mensaje = f"y{i+1} = {y}"
    print(mensaje)
    acum += y
    row.append(y)

while (Almacen < acum):
    acum = 0
    rows +=1
    alfa -= 0.1
    print(f"alfa={alfa}")
    for i, row in enumerate(data):
        y = round(math.sqrt((2 * row[1] * row[2]) / (row[3] - (2*alfa*row[4]))),2)
        print(f"K={row[1]}, D={row[2]}, H={row[3]}")
        mensaje = f"y{i+1} = {y}"
        print(mensaje)
        acum += y
        row.append(y)
    
print(tabulate(data))

nuevo_libro = openpyxl.Workbook()
hoja_calculo = nuevo_libro.active

# Escribir los encabezados en la hoja de cálculo
headers = ["Articulo", "K", "D", "H", "a",]
if rows>0:
    for r in range(rows+1):
        headers.append(f"Y{r+1}",)

for col_num, header in enumerate(headers, 1):
    col_letra = get_column_letter(col_num)
    celda = hoja_calculo[f"{col_letra}1"]
    celda.value = header
    celda.alignment = Alignment(horizontal='center')

# Escribir los datos en la hoja de cálculo
for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        col_letra = get_column_letter(col_num)
        celda = hoja_calculo[f"{col_letra}{row_num}"]
        celda.value = cell_value
        celda.alignment = Alignment(horizontal='center')

# Guardar el archivo Excel
nuevo_libro.save('Almacen_resultados.xlsx')